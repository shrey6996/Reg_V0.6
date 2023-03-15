import json
import os
import pyocr
import cv2
import secrets
import string
import shutil
import pyocr.builders
import pandas as pd
from PIL import Image
from werkzeug.utils import secure_filename
from flask import Flask, render_template, jsonify, request, send_file, redirect, session, send_from_directory, flash, \
    url_for
from flask_session import Session
import pandas as pd
from openpyxl import Workbook, load_workbook
from os.path import exists
from pdf2image import convert_from_path
from config import BaseFolder, UPLOAD_FOLDER, JsonPath

# BaseFolder = r"DocumentDataExtraction/"
# BaseFolder = r""
# UPLOAD_FOLDER = os.path.join(BaseFolder, "uploads")
# JsonPath = os.path.join(BaseFolder, 'info.json')
# form_fields = ["Name", "Invoice No.", "Invoice_Date", "Delivery_Note", "Mode_Terms_Of_Payment", "Reference_No_Date",
#                "Buyer's Order No.", "Delivery Note Date", "Dispatched through", "Destination", "Quantity",
#                "Total Amount"]

app = Flask(__name__)
# app.config['FormFields'] = form_fields
app.config['TESTING'] = True
app.secret_key = os.urandom(24)
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SECRET_KEY'] = os.urandom(24)
app.config['UPLOAD_EXTENSIONS'] = [".pdf", ".jpeg"]
app.config['JsonPath'] = JsonPath
app.config['UPLOAD_PATH'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['Extracted_Text'] = list()
Session(app)


############################ Create and Update Json config ###########################################################


# create json,Excel file, excel Sheet and then update the data in those files
def UpdateJson(CompanyName, InvoiceType, Data, Coordinates, JsonFilePath=app.config['JsonPath']):
    # Create a file if not Present
    mode = 'a' if exists(JsonFilePath) else 'w'
    with open(JsonFilePath, mode) as f:
        if mode == 'w':
            f.write("{}")

    with open(JsonFilePath, 'r+') as f:
        jsn = json.load(f)

        # #make new excel if company not in json or registering first time
        if CompanyName not in jsn:
            jsn[CompanyName] = {}
            with open(JsonFilePath, "w") as new_f:
                json.dump(jsn, new_f)

            wb = Workbook()
            wb.active
            wb.active.title = InvoiceType
            wb.save(filename=CompanyName + '.xlsx')
            print("The Excel Sheet has been Created in if loop for {} and {}".format(
                CompanyName, InvoiceType))

        elif InvoiceType not in jsn[CompanyName]:
            xl_path = CompanyName + ".xlsx"
            xl = load_workbook(xl_path)
            xl.create_sheet("{}".format(InvoiceType))
            xl.save(filename='{}.xlsx'.format(CompanyName))
            print("The Excel Sheet has been Created in elif loop for {} and {}".format(
                CompanyName, InvoiceType))

    jsn[CompanyName][InvoiceType] = Coordinates

    with open(JsonFilePath, "w") as new_f:
        json.dump(jsn, new_f)

    print("The json File has been updated with {} and {}".format(
        CompanyName, InvoiceType))
    UpdateExcelData(CompanyName, InvoiceType, Data)
    print("The Excel File has been updated with {} and {}".format(
        CompanyName, InvoiceType))


############################ Read Json config ###########################################################


def read_json_file(filename):
    f = open(filename)
    data = json.load(f)
    f.close()
    return data


############################ Update the Data In Excel ###########################################################


def UpdateExcelData(CompanyName, InvoiceType, Data):
    df = pd.DataFrame([Data])
    OldData = pd.read_excel("{}.xlsx".format(
        CompanyName), sheet_name=InvoiceType)
    NewData = pd.concat([df, OldData])
    NewData.reset_index(drop=True)
    xl_path = CompanyName + ".xlsx"
    with pd.ExcelWriter(xl_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        NewData.to_excel(writer, sheet_name=InvoiceType, index=False)


############################ tesseract config ###########################################################

#only for windows
# pyocr.tesseract.TESSERACT_CMD = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

tools = pyocr.get_available_tools()
if len(tools) == 0:
    print("No OCR tool found")
tool = tools[0]
langs = tool.get_available_languages()
lang = langs[0]


def get_text(image_path, r):
    img = cv2.imread(image_path)
    img_crop = img[int(r[1]):int(r[1] + r[3]), int(r[0]):int(r[0] + r[2])]

    new_path = r"uploads/{}text_image.jpg".format(session['FirstPageImageName'])
    cv2.imwrite(new_path, img_crop)
    txt = tool.image_to_string(
        Image.open(new_path),
        lang=lang,
        builder=pyocr.builders.TextBuilder()
    )
    os.remove(new_path)
    return txt


# main Registration page
@app.route('/', methods=['GET', 'POST'])
def main():
    if request.method == 'GET':
        if "FirstPage" in session:
            session.pop('FirstPage', None)
        if "NewRegistrationData" in session:
            session["NewRegistrationData"] = {}
            session['Data'] = {}
            session['Counter'] = 0
        return render_template('Index.html')

    if request.method == 'POST':
        uploaded_file = request.files['file1']
        filename = secure_filename(uploaded_file.filename)
        if filename != '':
            file_ext = os.path.splitext(filename)[1]
            if file_ext in app.config['UPLOAD_EXTENSIONS']:
                if file_ext == ".pdf":
                    uploaded_file.save(os.path.join(
                        app.config['UPLOAD_PATH'], filename))
                    file_path = os.path.join(app.config['UPLOAD_PATH'], filename)
                    session['FirstPageImageName'] = "{}.jpg".format(
                        ''.join(secrets.choice(string.ascii_uppercase + string.ascii_lowercase) for i in range(14)))
                    session['FirstPage'] = os.path.join(UPLOAD_FOLDER, session['FirstPageImageName'])
                    #windows
                    # pages = convert_from_path(
                    #     file_path, 200, poppler_path=r"C:\\Program Files\\poppler-0.68.0\\bin", fmt="jpeg")
                    
                    #linux
                    pages = convert_from_path(
                        file_path, 200, fmt="jpeg")
                    
                    pages[0].save(session['FirstPage'], 'JPEG')
                elif file_ext == ".jpeg":
                    session['FirstPageImageName'] = "{}.jpeg".format(
                        ''.join(secrets.choice(string.ascii_uppercase + string.ascii_lowercase) for i in range(14)))
                    uploaded_file.save(os.path.join(app.config['UPLOAD_PATH'], session['FirstPageImageName']))
                    session['FirstPage'] = os.path.join(UPLOAD_FOLDER, session['FirstPageImageName'])
                else:
                    return '<h5><p style="color:red">File Format Error, please Upload a PDF or Jpeg format file.</p> ' \
                           '</h5> <br> '
            else:
                return '<h5><p style="color:red">File Format Error, please Upload a PDF format file.</p> </h5> <br>'
        return '<h4><p style="color:green">File Uploaded!!</p> </h4> <br><div class="topnav"> <a class="active" ' \
               'href="/RegistrationProcess"> Next! </a> '


# Registration Process Crop functionality
@app.route('/RegistrationProcess', methods=['GET'])
def registration_process():
    return render_template('RegristrationProcess.html', filename=session['FirstPageImageName'])


@app.route('/display/<filename>')
def display_image(filename):
    return send_from_directory(app.config['UPLOAD_PATH'], filename)


# SendCoordinates
@app.route('/SendCoordinates', methods=['GET', 'POST'])
def SendCoordinates():
    if request.method == 'POST':
        data = request.get_json()
        # print(data['Snip_lableName'])
        # print(data['id'])
        # print(data["values"])
        Coordinates = data["values"]
        if not "NewRegistrationData" in session:
            session["NewRegistrationData"] = {}
            session['Data'] = {}
            session['Counter'] = 0
            session['NewRegistrationData'][data['Snip_lableName']] = [
                Coordinates['x'], Coordinates['y'], Coordinates['w'], Coordinates['h']]
        else:
            session['NewRegistrationData'][data['Snip_lableName']] = [
                Coordinates['x'], Coordinates['y'], Coordinates['w'], Coordinates['h']]

    extractedText = get_text(session['FirstPage'], session['NewRegistrationData'][data['Snip_lableName']])
    session['Data'][data['Snip_lableName']] = extractedText
    return {"value": extractedText}


@app.route('/RegistrationProcess', methods=['POST'])
def RegistrationData():
    if request.method == 'POST':
        FormData = request.form
        print("Form Data" , FormData)
        print("Session Data", session['Data'])
        Company_Name = request.form['CompanyName']
        Invoice_Type = request.form['InvoiceType']
        UpdateJson(CompanyName=Company_Name, InvoiceType=Invoice_Type,
                   Data=session['Data'], Coordinates=session['NewRegistrationData'],
                   JsonFilePath=app.config['JsonPath'])
        session.pop('Data', None)
        session.pop('NewRegistrationData', None)
        os.remove(session['FirstPage'])
        session.pop('FirstPage', None)
        session.pop('FirstPageImageName', None)
        flash('Registration completed for Company: "{}" and its Invoice Type: "{}"'.format(Company_Name, Invoice_Type))
    return redirect(url_for("main"))


# Second Page for uploading the Registered Files
@app.route('/Registered', methods=['GET'])
def Registered():
    f = open('info.json')
    data = json.load(f)
    f.close()
    OutputArray = []
    for row in data:
        outputObj = {
            'brand_id': row,
            'brand_name': row}
        OutputArray.append(outputObj)
    return render_template('RegisteredInvoiceExtraction.html', Company_Name=OutputArray)


# updates the invoice types
@app.route("/Registered/InvoiceType", methods=["POST", "GET"])
def InvoiceType():
    if request.method == 'POST':
        category_id = request.form['category_id']
        f = open('info.json')
        data = json.load(f)
        f.close()
        OutputArray = []
        for row in data[category_id]:
            outputObj = {
                'id': category_id,
                'name': row}
            OutputArray.append(outputObj)
        print(OutputArray)
    return jsonify(OutputArray)


# accepts the PDF files from the user
@app.route('/UploadInvoice', methods=["POST", "GET"])
def UploadInvoice():
    if request.method == 'POST':
        if request.form.get('Company_Name') and request.form.get('Invoice_Type'):
            Company_Name = request.form.get('Company_Name')
            session['Company_Name'] = Company_Name
            Invoice_Type = request.form.get('Invoice_Type')
            data = read_json_file(app.config['JsonPath'])
        else:
            return "Select the Company Name and Invoice Type  then try to upload your file.", 400
        if request.files['file']:
            uploaded_file = request.files['file']
            filename = secure_filename(uploaded_file.filename)
            if filename != '':
                print("File Name is not empty")
                file_ext = os.path.splitext(filename)[1]
                if file_ext in app.config['UPLOAD_EXTENSIONS']:
                    uploaded_file.save(os.path.join(app.config['UPLOAD_PATH'], filename))
                    file_path = os.path.join(app.config['UPLOAD_PATH'], filename)
                    session['FirstPageImageName'] = "{}.jpeg".format(
                        ''.join(secrets.choice(string.ascii_uppercase + string.ascii_lowercase) for i in range(14)))
                    if file_ext == ".pdf":
                        print("PDF file")
                        FirstPageSavePath = os.path.join(app.config['UPLOAD_PATH'], session['FirstPageImageName'])
                        #windows
                        # pages = convert_from_path(
                        #     file_path, 200, poppler_path=r"C:\\Program Files\\poppler-0.68.0\\bin", fmt="jpeg")
                        
                        #linux
                        pages = convert_from_path(
                            file_path, 200, fmt="jpeg")

                        pages[0].save(FirstPageSavePath, 'JPEG')
                    elif file_ext == ".jpeg":
                        print("JPEG file")
                        FirstPageSavePath = os.path.join(app.config['UPLOAD_PATH'], session['FirstPageImageName'])
                        shutil.copyfile(file_path, FirstPageSavePath)
                    else:
                        return '<h5><p style="color:red">File Format Error, please Upload a PDF or Jpeg format ' \
                               'file.</p> </h5> <br> '
                else:
                    return '<h5><p style="color:red">File Format Error, please Upload a PDF format file.</p> </h5> <br>', 400
        TextData = {}
        try:
            print(FirstPageSavePath)
            for i in data[Company_Name][Invoice_Type]:
                print(i)
                print(data[Company_Name][Invoice_Type][i][0])
                textFromImage = get_text(FirstPageSavePath, (data[Company_Name][Invoice_Type][i][0], data[Company_Name]
                [Invoice_Type][i][1], data[Company_Name][Invoice_Type][i][2], data[Company_Name][Invoice_Type][i][3]))
                new_dict = {str(i): str(textFromImage)}
                TextData.update(new_dict)
            print(TextData)
            UpdateExcelData(CompanyName=Company_Name,
                            InvoiceType=Invoice_Type, Data=TextData)
        except:
            return "Got some error Please try again for pdf : {}".format(filename), 400
        os.remove(file_path)
        os.remove(FirstPageSavePath)
        session.pop('FirstPageImageName', None)
    return 'Success', 204


# Download function for the xlsx file using the session name of the company
@app.route('/Downloads', methods=['GET', 'POST'])
def Downloads():
    file_name = '{}.xlsx'.format(session['Company_Name'])
    return send_file(file_name, as_attachment=True)


# main driving function
if __name__ == '__main__':
    app.run(port=4224, debug=True, host = '0.0.0.0')
